# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# MIT License. See license.txt
import re
from io import BytesIO

import openpyxl
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import frappe

# AGREGADO POR SEBAS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
# FIN AGREGADO

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')


# return xlsx file object
def make_xlsx(data, sheet_name, wb=None, column_widths=None):
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook(write_only=True)

	ws = wb.create_sheet(sheet_name, 0)

	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	row1 = ws.row_dimensions[1]
	row1.font = Font(name='Calibri', bold=True)

	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, str) and (sheet_name not in ['Data Import Template', 'Data Export']):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = re.sub(ILLEGAL_CHARACTERS_RE, '', value)

			clean_row.append(value)

		ws.append(clean_row)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file


def handle_html(data):
	# return if no html tags found
	data = frappe.as_unicode(data)

	if '<' not in data:
		return data
	if '>' not in data:
		return data

	from html2text import HTML2Text

	h = HTML2Text()
	h.unicode_snob = True
	h = h.unescape(data or "")

	obj = HTML2Text()
	obj.ignore_links = True
	obj.body_width = 0

	try:
		value = obj.handle(h)
	except Exception:
		# unable to parse html, send it raw
		return data

	value = ", ".join(value.split('  \n'))
	value = " ".join(value.split('\n'))
	value = ", ".join(value.split('# '))

	return value


def read_xlsx_file_from_attached_file(file_url=None, fcontent=None, filepath=None):
	if file_url:
		_file = frappe.get_doc("File", {"file_url": file_url})
		filename = _file.get_full_path()
	elif fcontent:
		filename = BytesIO(fcontent)
	elif filepath:
		filename = filepath
	else:
		return

	rows = []
	wb1 = load_workbook(filename=filename, read_only=True, data_only=True)
	ws1 = wb1.active
	for row in ws1.iter_rows():
		tmp_list = []
		for cell in row:
			tmp_list.append(cell.value)
		rows.append(tmp_list)
	return rows


def read_xls_file_from_attached_file(content):
	book = xlrd.open_workbook(file_contents=content)
	sheets = book.sheets()
	sheet = sheets[0]
	rows = []
	for i in range(sheet.nrows):
		rows.append(sheet.row_values(i))
	return rows


def build_xlsx_response(data, filename):
	xlsx_file = make_xlsx(data, filename)
	# write out response as a xlsx type
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def make_xlsx_sebas(data, sheet_name, wb=None, column_widths=None):
	column_widths = column_widths or []
	if wb is None:
		wb = Workbook()

	ws = wb.active
	ws.title = sheet_name

	# Define the style for the header
	header_font = Font(name='Calibri', bold=True, color='FFFFFF')  # White text
	header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')  # Black background

	# Dictionary to store max length for each column
	max_lengths = {}

	for row_index, row in enumerate(data):
		if not isinstance(row, (list, tuple)):
			raise TypeError(f"Invalid row type: {type(row)}. Each row must be a list or tuple.")

		# Apply header style
		if row_index == 0:  # First row (header)
			for col_index, cell_value in enumerate(row):
				cell = ws.cell(row=row_index + 1, column=col_index + 1, value=cell_value)
				cell.font = header_font
				cell.fill = header_fill
		else:
			ws.append(row)

		# Update max lengths
		for col_index, cell_value in enumerate(row):
			cell_length = len(str(cell_value))
			if col_index not in max_lengths:
				max_lengths[col_index] = cell_length
			else:
				max_lengths[col_index] = max(max_lengths[col_index], cell_length)

	# Set column widths based on max lengths
	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	# Adjust column widths dynamically if no specific width is provided
	for col_index, length in max_lengths.items():
		col_letter = get_column_letter(col_index + 1)
		# Set width with a slight buffer
		ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 50)  # Adjust buffer and max width as needed

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response_sebas(data, filename):
	xlsx_file = make_xlsx_sebas(data, filename)
	# write out response as a xlsx type
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'